import pandas as pd
from typing import List, Tuple
from icecream import ic
import openpyxl
from loader import rq
from utils.parser import Parse


class Excel_db:
    file_name = 'db.xlsx'

    @classmethod
    def create_xl(cls, cnx):
        df = pd.read_sql_query("SELECT * FROM cards", cnx)
        df.to_excel(cls.file_name, index=False, startcol=1, startrow=1)
        cls._change_width()

    @classmethod
    def _change_width(cls):
        wb = openpyxl.load_workbook(cls.file_name)
        sheet = wb.active

        for col_idx in range(1, sheet.max_column + 1):
            max_length = 0
            for cell in sheet.iter_cols(min_col=col_idx, max_col=col_idx):
                for cell_obj in cell:
                    if len(str(cell_obj.value)) > max_length:
                        max_length = len(str(cell_obj.value))
            sheet.column_dimensions[openpyxl.utils.get_column_letter(
                col_idx)].width = max_length + 2

            # Выравнивание колонки по тексту
            for cell in sheet.iter_cols(min_col=col_idx, max_col=col_idx):
                for cell_obj in cell:
                    cell_obj.alignment = openpyxl.styles.Alignment(
                        horizontal='left')

        wb.save(cls.file_name)


class Excel_read():
    @classmethod
    async def read(cls, file_name):
        df = pd.read_excel(file_name, index_col=1, skiprows=1, skipfooter=1)
        # df = df[["status","id_telegram"]]
        # создаем новый dataframe where status == 4
        status_del_df = df[df['status'] == 4]
        index_list_del = status_del_df.index.tolist()  # получаем список индексов
        with rq as cursor:
            for id in index_list_del:
                ic(id)
                cursor.execute(f"DELETE FROM cards WHERE id = {id}")
        # создаем новый dataframe where status == 3
        new_df = df[df['status'] == 3]
        index_list = new_df.index.tolist()  # получаем список индексов
        for id in index_list:
            with rq:
                database_address = rq.select_one(
                    "cards", ["address"], f"id = {id}")[0]
            address = new_df.loc[id, 'address']
            company_name = new_df.loc[id, 'company_name']
            comment = new_df.loc[id, 'comment']
            coordinates = new_df.loc[id, 'coordinates']
            ic(database_address, address)
            coordinates, address = await Parse.parse_coordinates_and_address(address)
            if coordinates is None or coordinates == "---":
                # ic(coordinates, "1")
                coordinates = "---"
                status = 2
            else:
                # ic(coordinates)
                status = 1
            with rq:
                rq.write_update("cards", [("company_name", company_name), ("address", address), (
                    "comment", comment), ("status", status), ("coordinates", coordinates)], f'id = {id}')


if __name__ == "__main__":
    # import sqlite3
    # cnx = sqlite3.connect("request.db")
    Excel_read().read()
